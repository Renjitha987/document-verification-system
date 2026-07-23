import io
import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse, Http404
from django.core.mail import EmailMessage
from django.conf import settings
from django.utils import timezone
from django.core.paginator import Paginator

import openpyxl

from accounts.models import User, StudentProfile
from .models import Department, Course, Certificate
from .forms import DepartmentForm, CourseForm, CertificateForm
from .utils import generate_verification_pdf
from audit_logs.models import AuditLog, VerificationHistory
from audit_logs.utils import log_action

# Helper to check admin role
def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_admin():
            messages.error(request, "Access denied. Administrator privileges required.")
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
@admin_required
def admin_dashboard(request):
    """Generates statistics and Chart.js datasets for the admin portal dashboard."""
    # Summary stats
    total_students = StudentProfile.objects.count()
    total_certificates = Certificate.objects.count()
    
    verified_certs = Certificate.objects.filter(status='APPROVED').count()
    pending_certs = Certificate.objects.filter(status='PENDING').count()
    revoked_certs = Certificate.objects.filter(status='REVOKED').count()
    
    # Today's verifications
    today = timezone.now().date()
    today_verifications = VerificationHistory.objects.filter(
        timestamp__date=today, status='SUCCESS'
    ).count()
    
    # Monthly verifications count (current month)
    current_month = timezone.now().month
    current_year = timezone.now().year
    monthly_verifications = VerificationHistory.objects.filter(
        timestamp__month=current_month,
        timestamp__year=current_year,
        status='SUCCESS'
    ).count()
    
    # Active admin/student users
    active_users = User.objects.filter(is_active=True).count()
    
    # Recent items
    recent_activities = AuditLog.objects.select_related('user').all()[:6]
    recent_verifications = VerificationHistory.objects.all()[:6]
    
    # Chart 1: Status Distribution (Doughnut)
    status_distribution = list(Certificate.objects.values('status').annotate(count=Count('status')))
    
    # Chart 2: Department Wise Certificates (Pie/Doughnut)
    dept_distribution = list(Certificate.objects.values('department__code').annotate(count=Count('id')))
    
    # Chart 3: Student Distribution by Department (Bar)
    student_distribution = list(StudentProfile.objects.values('department__code').annotate(count=Count('id')))
    
    # Chart 4: Monthly Certificates (past 6 months)
    # Basic aggregate: group by issue_date month
    monthly_certs = []
    for i in range(5, -1, -1):
        check_date = timezone.now() - datetime.timedelta(days=i*30)
        month_name = check_date.strftime('%B')
        count = Certificate.objects.filter(
            issue_date__month=check_date.month,
            issue_date__year=check_date.year
        ).count()
        monthly_certs.append({'month': month_name, 'count': count})
        
    context = {
        'total_students': total_students,
        'total_certificates': total_certificates,
        'verified_certs': verified_certs,
        'pending_certs': pending_certs,
        'revoked_certs': revoked_certs,
        'today_verifications': today_verifications,
        'monthly_verifications': monthly_verifications,
        'active_users': active_users,
        'recent_activities': recent_activities,
        'recent_verifications': recent_verifications,
        
        # JSON data for Chart.js
        'status_dist': status_distribution,
        'dept_dist': dept_distribution,
        'student_dist': student_distribution,
        'monthly_certs': monthly_certs,
    }
    return render(request, 'certificates/admin_dashboard.html', context)


# DEPARTMENTS CRUD
@login_required
@admin_required
def manage_departments(request):
    departments = Department.objects.all()
    form = DepartmentForm()
    return render(request, 'certificates/departments.html', {'departments': departments, 'form': form})


@login_required
@admin_required
def add_department(request):
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            dept = form.save()
            log_action(request.user, 'ADD_DEPARTMENT', f"Created department {dept.name} ({dept.code}).", request)
            messages.success(request, f"Department '{dept.name}' added successfully!")
        else:
            messages.error(request, "Failed to add department. Code must be unique.")
    return redirect('manage_departments')


@login_required
@admin_required
def delete_department(request, pk):
    dept = get_object_or_404(Department, pk=pk)
    try:
        dept.delete()
        log_action(request.user, 'DELETE_DEPARTMENT', f"Deleted department {dept.name}.", request)
        messages.success(request, f"Department '{dept.name}' deleted successfully!")
    except Exception:
        messages.error(request, f"Cannot delete department '{dept.name}' because it contains associated records.")
    return redirect('manage_departments')


# COURSES CRUD
@login_required
@admin_required
def manage_courses(request):
    courses = Course.objects.select_related('department').all()
    form = CourseForm()
    return render(request, 'certificates/courses.html', {'courses': courses, 'form': form})


@login_required
@admin_required
def add_course(request):
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            course = form.save()
            log_action(request.user, 'ADD_COURSE', f"Created course {course.name} ({course.code}).", request)
            messages.success(request, f"Course '{course.name}' added successfully!")
        else:
            messages.error(request, "Failed to add course. Code must be unique.")
    return redirect('manage_courses')


@login_required
@admin_required
def delete_course(request, pk):
    course = get_object_or_404(Course, pk=pk)
    try:
        course.delete()
        log_action(request.user, 'DELETE_COURSE', f"Deleted course {course.name}.", request)
        messages.success(request, f"Course '{course.name}' deleted successfully!")
    except Exception:
        messages.error(request, f"Cannot delete course '{course.name}' because it contains associated records.")
    return redirect('manage_courses')


# STUDENTS CRUD
@login_required
@admin_required
def manage_students(request):
    query = request.GET.get('q', '').strip()
    students_list = StudentProfile.objects.select_related('user', 'course', 'department').all()
    
    if query:
        students_list = students_list.filter(
            Q(register_number__icontains=query) |
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(user__email__icontains=query) |
            Q(course__name__icontains=query) |
            Q(department__name__icontains=query)
        )
        
    paginator = Paginator(students_list, 15)
    page_number = request.GET.get('page')
    students = paginator.get_page(page_number)
    
    departments = Department.objects.all()
    courses = Course.objects.all()
    
    context = {
        'students': students,
        'query': query,
        'departments': departments,
        'courses': courses,
    }
    return render(request, 'certificates/students.html', context)


@login_required
@admin_required
@transaction.atomic
def add_student(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        
        register_number = request.POST.get('register_number', '').strip()
        father_name = request.POST.get('father_name', '').strip()
        mother_name = request.POST.get('mother_name', '').strip()
        department_id = request.POST.get('department', '')
        course_id = request.POST.get('course', '')
        college_name = request.POST.get('college_name', '').strip()
        university_name = request.POST.get('university_name', '').strip()
        
        # Simple backend checks
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect('manage_students')
        if StudentProfile.objects.filter(register_number=register_number).exists():
            messages.error(request, "Register Number already exists.")
            return redirect('manage_students')
            
        try:
            # Create user
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                role='student'
            )
            
            # Fetch ForeignKey objects
            dept = Department.objects.get(id=department_id)
            course = Course.objects.get(id=course_id)
            
            # Create profile
            profile = StudentProfile.objects.create(
                user=user,
                register_number=register_number,
                father_name=father_name,
                mother_name=mother_name,
                department=dept,
                course=course,
                college_name=college_name,
                university_name=university_name
            )
            
            log_action(request.user, 'ADD_STUDENT', f"Created student user {username} with register number {register_number}.", request)
            messages.success(request, f"Student '{first_name} {last_name}' added successfully!")
            
        except Exception as e:
            messages.error(request, f"Error saving student: {str(e)}")
            
    return redirect('manage_students')


@login_required
@admin_required
@transaction.atomic
def edit_student(request, pk):
    profile = get_object_or_404(StudentProfile, pk=pk)
    user = profile.user
    
    if request.method == 'POST':
        user.first_name = request.POST.get('first_name', '').strip()
        user.last_name = request.POST.get('last_name', '').strip()
        user.email = request.POST.get('email', '').strip()
        if request.POST.get('password', '').strip():
            user.set_password(request.POST.get('password'))
        user.save()
        
        profile.father_name = request.POST.get('father_name', '').strip()
        profile.mother_name = request.POST.get('mother_name', '').strip()
        profile.department = Department.objects.get(id=request.POST.get('department'))
        profile.course = Course.objects.get(id=request.POST.get('course'))
        profile.college_name = request.POST.get('college_name', '').strip()
        profile.university_name = request.POST.get('university_name', '').strip()
        profile.save()
        
        log_action(request.user, 'EDIT_STUDENT', f"Modified student profile {profile.register_number}.", request)
        messages.success(request, f"Student profile '{profile.register_number}' updated successfully.")
        return redirect('manage_students')
        
    departments = Department.objects.all()
    courses = Course.objects.all()
    context = {
        'profile': profile,
        'departments': departments,
        'courses': courses,
    }
    return render(request, 'certificates/edit_student.html', context)


@login_required
@admin_required
@transaction.atomic
def delete_student(request, pk):
    profile = get_object_or_404(StudentProfile, pk=pk)
    user = profile.user
    try:
        user.delete() # Cascade deletes the profile
        log_action(request.user, 'DELETE_STUDENT', f"Deleted student with register number {profile.register_number}.", request)
        messages.success(request, f"Student '{profile.register_number}' deleted successfully.")
    except Exception as e:
        messages.error(request, f"Error deleting student: {str(e)}")
    return redirect('manage_students')


# CERTIFICATES CRUD
def get_filtered_certificates(request):
    """Helper to query certificates based on GET parameters filter."""
    cert_num = request.GET.get('cert_num', '').strip()
    reg_num = request.GET.get('reg_num', '').strip()
    student_name = request.GET.get('student_name', '').strip()
    dept_id = request.GET.get('department', '').strip()
    course_id = request.GET.get('course', '').strip()
    year = request.GET.get('year', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    status = request.GET.get('status', '').strip()

    qs = Certificate.objects.select_related('student__user', 'course', 'department').all()

    if cert_num:
        qs = qs.filter(certificate_number__icontains=cert_num)
    if reg_num:
        qs = qs.filter(student__register_number__icontains=reg_num)
    if student_name:
        qs = qs.filter(student_name_snapshot__icontains=student_name)
    if dept_id:
        qs = qs.filter(department_id=dept_id)
    if course_id:
        qs = qs.filter(course_id=course_id)
    if year:
        qs = qs.filter(issue_date__year=year)
    if start_date:
        qs = qs.filter(issue_date__gte=start_date)
    if end_date:
        qs = qs.filter(issue_date__lte=end_date)
    if status:
        qs = qs.filter(status=status)
        
    return qs.order_by('-created_at')


@login_required
@admin_required
def manage_certificates(request):
    certificates_qs = get_filtered_certificates(request)
    
    paginator = Paginator(certificates_qs, 15)
    page_number = request.GET.get('page')
    certificates = paginator.get_page(page_number)
    
    departments = Department.objects.all()
    courses = Course.objects.all()
    
    context = {
        'certificates': certificates,
        'departments': departments,
        'courses': courses,
        # Preserve search queries in template forms
        'filters': request.GET,
    }
    return render(request, 'certificates/certificates.html', context)


@login_required
@admin_required
def add_certificate(request):
    if request.method == 'POST':
        form = CertificateForm(request.POST, request.FILES)
        if form.is_valid():
            cert = form.save()
            log_action(request.user, 'ADD_CERTIFICATE', f"Added certificate {cert.certificate_number} for student profile {cert.student.register_number}.", request)
            messages.success(request, f"Certificate '{cert.certificate_number}' created successfully!")
            return redirect('manage_certificates')
        else:
            messages.error(request, "Failed to create certificate. Ensure certificate number is unique if manually entered.")
    else:
        form = CertificateForm()
        
    return render(request, 'certificates/add_certificate.html', {'form': form})


@login_required
@admin_required
def edit_certificate(request, pk):
    cert = get_object_or_404(Certificate, pk=pk)
    if request.method == 'POST':
        form = CertificateForm(request.POST, request.FILES, instance=cert)
        if form.is_valid():
            cert = form.save()
            log_action(request.user, 'EDIT_CERTIFICATE', f"Updated certificate {cert.certificate_number}.", request)
            messages.success(request, f"Certificate '{cert.certificate_number}' updated successfully.")
            return redirect('manage_certificates')
    else:
        form = CertificateForm(instance=cert)
        
    return render(request, 'certificates/edit_certificate.html', {'form': form, 'cert': cert})


@login_required
@admin_required
def delete_certificate(request, pk):
    cert = get_object_or_404(Certificate, pk=pk)
    cert_num = cert.certificate_number
    cert.delete()
    log_action(request.user, 'DELETE_CERTIFICATE', f"Deleted certificate {cert_num}.", request)
    messages.success(request, f"Certificate '{cert_num}' deleted successfully.")
    return redirect('manage_certificates')


@login_required
@admin_required
def update_certificate_status(request, pk, status_choice):
    cert = get_object_or_404(Certificate, pk=pk)
    
    if status_choice not in ['PENDING', 'APPROVED', 'REJECTED', 'REVOKED']:
        messages.error(request, "Invalid status choice.")
        return redirect('manage_certificates')
        
    cert.status = status_choice
    if status_choice == 'APPROVED':
        cert.verification_date = timezone.now()
    else:
        cert.verification_date = None
        
    cert.save(update_fields=['status', 'verification_date'])
    log_action(request.user, 'UPDATE_STATUS', f"Changed status of certificate {cert.certificate_number} to {status_choice}.", request)
    messages.success(request, f"Status of certificate '{cert.certificate_number}' updated to {cert.get_status_display()}.")
    return redirect('manage_certificates')


# EXPORTS
@login_required
@admin_required
def export_certificates_excel(request):
    """Exports the filtered list of certificates to a clean Excel spreadsheet."""
    certificates = get_filtered_certificates(request)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Certificate Registry"
    
    # Define headers
    headers = [
        "Certificate ID", "Certificate Number", "Register Number", 
        "Student Name", "Father Name", "Mother Name", 
        "Course", "Department", "CGPA", "Grade", 
        "Issue Date", "Status", "Digital Signature"
    ]
    ws.append(headers)
    
    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        
    # Append rows
    for c in certificates:
        ws.append([
            str(c.id),
            c.certificate_number,
            c.student.register_number,
            c.student_name_snapshot,
            c.student.father_name,
            c.student.mother_name,
            c.course.code,
            c.department.code,
            float(c.cgpa),
            c.grade,
            c.issue_date.strftime('%Y-%m-%d'),
            c.get_status_display(),
            c.digital_signature
        ])
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="certificate_registry_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    
    log_action(request.user, 'EXPORT_EXCEL', "Exported certificate registry to Excel.", request)
    return response


@login_required
@admin_required
def export_certificates_pdf(request):
    """Generates a PDF registry report containing a table of the selected certificates."""
    certificates = get_filtered_certificates(request)
    
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="certificate_registry_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    # Use landscape mode for tables with many columns
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'RegistryTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#0f172a'),
        alignment=1, # Center
        spaceAfter=15
    )
    
    th_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white
    )
    
    tb_style = ParagraphStyle(
        'TableBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#0f172a')
    )
    
    story.append(Paragraph("Smart Certificate Verification Portal", title_style))
    story.append(Paragraph(f"Active Certificate Registry &bull; Generated on {timezone.now().strftime('%Y-%m-%d %H:%M:%S UTC')}", ParagraphStyle('Sub', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=12, alignment=1, spaceAfter=20)))
    
    # Build Table
    table_data = [[
        Paragraph("<b>No.</b>", th_style),
        Paragraph("<b>Cert Number</b>", th_style),
        Paragraph("<b>Reg Number</b>", th_style),
        Paragraph("<b>Student Name</b>", th_style),
        Paragraph("<b>Course</b>", th_style),
        Paragraph("<b>Department</b>", th_style),
        Paragraph("<b>CGPA</b>", th_style),
        Paragraph("<b>Issue Date</b>", th_style),
        Paragraph("<b>Status</b>", th_style),
    ]]
    
    for idx, c in enumerate(certificates, 1):
        table_data.append([
            Paragraph(str(idx), tb_style),
            Paragraph(c.certificate_number, tb_style),
            Paragraph(c.student.register_number, tb_style),
            Paragraph(c.student_name_snapshot, tb_style),
            Paragraph(c.course.code, tb_style),
            Paragraph(c.department.code, tb_style),
            Paragraph(f"{c.cgpa} ({c.grade})", tb_style),
            Paragraph(c.issue_date.strftime('%Y-%m-%d'), tb_style),
            Paragraph(c.get_status_display(), tb_style),
        ])
        
    table = Table(table_data, colWidths=[0.4*inch, 1.3*inch, 1.2*inch, 1.8*inch, 1.1*inch, 1.0*inch, 0.8*inch, 1.0*inch, 1.0*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e293b')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    
    story.append(table)
    doc.build(story)
    
    log_action(request.user, 'EXPORT_PDF', "Exported certificate registry report to PDF.", request)
    return response


# EMAIL DISPATCHER
@login_required
@admin_required
def email_verification_report(request, pk):
    """Sends the verification report PDF via SMTP email to the student user email or custom email."""
    cert = get_object_or_404(Certificate.objects.select_related('student__user', 'course', 'department'), pk=pk)
    
    recipient_email = request.POST.get('custom_email', '').strip()
    if not recipient_email:
        recipient_email = cert.student.user.email
        
    if not recipient_email:
        messages.error(request, "Recipient email address not found. Please provide an email address.")
        return redirect('manage_certificates')
        
    try:
        # Generate the PDF in-memory
        pdf_buffer = generate_verification_pdf(cert)
        pdf_data = pdf_buffer.getvalue()
        
        # Build email content
        subject = f"[Smart Portal] Certificate Verification Report: {cert.certificate_number}"
        body = (
            f"Dear Recipient,\n\n"
            f"Please find attached the official verification report for Certificate Number: {cert.certificate_number}.\n\n"
            f"Record Metadata:\n"
            f"- Student Name: {cert.student_name_snapshot}\n"
            f"- Register Number: {cert.student.register_number}\n"
            f"- Course: {cert.course.name}\n"
            f"- Department: {cert.department.name}\n"
            f"- Verification Status: {cert.get_status_display()}\n\n"
            f"This is an automated delivery. You can verify this certificate online at any time by scanning the attached QR code.\n\n"
            f"Best regards,\n"
            f"Smart Certificate Verification Team"
        )
        
        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=settings.DEFAULT_FROM_EMAIL or 'portal@verification-system.org',
            to=[recipient_email]
        )
        
        # Attach the file
        email.attach(f"Verification_Report_{cert.certificate_number}.pdf", pdf_data, 'application/pdf')
        
        # Dispatch
        email.send(fail_silently=False)
        
        log_action(request.user, 'EMAIL_REPORT', f"Sent verification email for certificate {cert.certificate_number} to {recipient_email}.", request)
        messages.success(request, f"Verification report for '{cert.certificate_number}' successfully emailed to {recipient_email}!")
        
    except Exception as e:
        messages.error(request, f"SMTP Error: Failed to send email. Details: {str(e)}")
        
    return redirect('manage_certificates')
